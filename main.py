#!/usr/bin/env python3
"""
Automatizare Zile de Naștere & Onomastici — parohie.

Citește ZILE.xlsx, calculează cine are ziua de naștere / onomastica
MÂINE și PESTE O SĂPTĂMÂNĂ și trimite UN SINGUR mesaj WhatsApp combinat
prin Meta WhatsApp Cloud API.

Rulare normală (trimite mesaj real dacă e ceva de anunțat):
    python main.py

Testare fără a trimite nimic (doar printează mesajul):
    python main.py --dry-run
    python main.py --dry-run --date 2026-04-04
"""

import argparse
import datetime
import os
import sys

import openpyxl
import requests
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Configurare fixă (onomastici) — NU schimba fără confirmare.
# ---------------------------------------------------------------------------

# cheie = token normalizat (UPPERCASE, fără diacritice/spații) găsit în MENȚIUNI
# valoare = (luna, ziua)
FIXED_SAINTS = {
    "VASILE": (1, 1),       # 1 ianuarie
    "ION": (1, 7),          # 7 ianuarie
    "GABRIEL": (3, 26),     # 26 martie
    "GHEORGHE": (4, 23),    # 23 aprilie
    "CONSTANTIN": (5, 21),  # 21 mai
    "NICOLAE": (12, 6),     # 6 decembrie
    "MARIA": (8, 15),       # 15 august
}

# text de afișare (frumos) pentru fiecare token recunoscut
SAINT_DISPLAY = {
    "VASILE": "Sf. Vasile",
    "ION": "Sf. Ion",
    "GABRIEL": "Sf. Gabriel",
    "GHEORGHE": "Sf. Gheorghe",
    "CONSTANTIN": "Sf. Constantin",
    "NICOLAE": "Sf. Nicolae",
    "MARIA": "Sf. Maria",
    "FLORII": "Florii",
}

FLORII_TOKEN = "FLORII"

WHATSAPP_API_VERSION = "v20.0"

# ---------------------------------------------------------------------------
# Calcul Paștele Ortodox / Florii (dată mobilă)
# ---------------------------------------------------------------------------


def orthodox_easter(year: int) -> datetime.date:
    """Algoritmul Meeus (iulian) + offset de 13 zile -> dată gregoriană."""
    a = year % 4
    b = year % 7
    c = year % 19
    d = (19 * c + 15) % 30
    e = (2 * a + 4 * b - d + 34) % 7
    month = (d + e + 114) // 31
    day = ((d + e + 114) % 31) + 1
    julian_easter = datetime.date(year, month, day)
    return julian_easter + datetime.timedelta(days=13)


def florii_date(year: int) -> datetime.date:
    return orthodox_easter(year) - datetime.timedelta(days=7)


# ---------------------------------------------------------------------------
# Citire Excel
# ---------------------------------------------------------------------------


def _normalize_token(text: str) -> str:
    """Uppercase, fără diacritice, fără spații în plus."""
    if text is None:
        return ""
    text = str(text).strip().upper()
    replacements = {
        "Ă": "A", "Â": "A", "Î": "I", "Ș": "S", "Ş": "S", "Ț": "T", "Ţ": "T",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def _parse_birth_date(value) -> datetime.date | None:
    """Convertește valoarea din coloana C (DATA NAȘTERII) într-un datetime.date.

    Acceptă: datetime.datetime / datetime.date (cum dă de obicei openpyxl
    pentru celule formatate ca dată), string în formate uzuale, sau serial
    Excel (float/int), dacă celula nu a fost formatată ca dată.
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)):
        # serial Excel (epoch 1899-12-30, cu bug-ul de 1900 leap year inclus)
        try:
            return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(value)))
        except (OverflowError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%y"):
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None
    return None


class Person:
    __slots__ = ("nume", "prenume", "birth_date", "domiciliu", "mentiuni_raw", "mentiuni_tokens")

    def __init__(self, nume, prenume, birth_date, domiciliu, mentiuni_raw):
        self.nume = (nume or "").strip()
        self.prenume = (prenume or "").strip()
        self.birth_date = birth_date
        self.domiciliu = (domiciliu or "").strip()
        self.mentiuni_raw = (mentiuni_raw or "").strip()
        self.mentiuni_tokens = [
            t.strip() for t in self.mentiuni_raw.split(",") if t.strip()
        ] if self.mentiuni_raw else []

    @property
    def full_name(self) -> str:
        return f"{self.nume} {self.prenume}".strip()


def read_people(path: str) -> list[Person]:
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Nu găsesc fișierul Excel '{path}'. Pune ZILE.xlsx în directorul "
            "proiectului (sau indică-l cu --file / variabila EXCEL_FILE_PATH)."
        )

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
    else:
        ws = wb.active  # fallback dacă foaia are alt nume

    people: list[Person] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        # A, B, C, D(luna, ignorat), E, F
        nume = row[0] if len(row) > 0 else None
        prenume = row[1] if len(row) > 1 else None
        data_nasterii_raw = row[2] if len(row) > 2 else None
        domiciliu = row[4] if len(row) > 4 else None
        mentiuni = row[5] if len(row) > 5 else None

        if nume is None and prenume is None and data_nasterii_raw is None:
            continue  # rând complet gol

        birth_date = _parse_birth_date(data_nasterii_raw)
        if birth_date is None:
            continue  # ignorăm rândurile fără dată de naștere validă

        people.append(Person(nume, prenume, birth_date, domiciliu, mentiuni))

    return people


# ---------------------------------------------------------------------------
# Calcul date țintă (următoarea aniversare / onomastică)
# ---------------------------------------------------------------------------


def _safe_date(year: int, month: int, day: int) -> datetime.date:
    """Construiește o dată, cu fallback pentru 29 februarie în an nebisect."""
    try:
        return datetime.date(year, month, day)
    except ValueError:
        if month == 2 and day == 29:
            return datetime.date(year, 2, 28)
        raise


def next_occurrence(month: int, day: int, today: datetime.date) -> datetime.date:
    """Următoarea apariție a unei date (lună/zi) începând cu azi (inclusiv azi)."""
    candidate = _safe_date(today.year, month, day)
    if candidate < today:
        candidate = _safe_date(today.year + 1, month, day)
    return candidate


def next_florii(today: datetime.date) -> datetime.date:
    candidate = florii_date(today.year)
    if candidate < today:
        candidate = florii_date(today.year + 1)
    return candidate


# ---------------------------------------------------------------------------
# Construirea raportului
# ---------------------------------------------------------------------------


class DayReport:
    def __init__(self, label: str, target_date: datetime.date):
        self.label = label
        self.target_date = target_date
        self.birthdays: list[str] = []   # linii formatate 🎂
        self.namedays: list[str] = []    # linii formatate ⛪

    @property
    def is_empty(self) -> bool:
        return not self.birthdays and not self.namedays


def _mentiuni_display(person: Person) -> str | None:
    """Text de afișare pentru MENȚIUNI, folosit lângă intrarea de zi de naștere."""
    if not person.mentiuni_tokens:
        return None
    parts = []
    for token in person.mentiuni_tokens:
        norm = _normalize_token(token)
        parts.append(SAINT_DISPLAY.get(norm, token.strip().title()))
    return ", ".join(parts)


def build_reports(people: list[Person], today: datetime.date) -> tuple[DayReport, DayReport]:
    tomorrow = today + datetime.timedelta(days=1)
    next_week = today + datetime.timedelta(days=7)

    report_tomorrow = DayReport("MÂINE", tomorrow)
    report_nextweek = DayReport("PESTE O SĂPTĂMÂNĂ", next_week)

    for target_date, report in ((tomorrow, report_tomorrow), (next_week, report_nextweek)):
        birthday_lines = []
        nameday_lines = []

        for person in people:
            # --- Zi de naștere ---
            bd = person.birth_date
            next_bday = next_occurrence(bd.month, bd.day, today)
            if next_bday == target_date:
                age = next_bday.year - bd.year
                mentiuni_txt = _mentiuni_display(person)
                line = (
                    f"🎂 {person.nume} {person.prenume} — zi de naștere, "
                    f"împlinește {age} ani"
                )
                if mentiuni_txt:
                    line += f" | Onomastică: {mentiuni_txt}"
                if person.domiciliu:
                    line += f" | Domiciliu: {person.domiciliu}"
                birthday_lines.append((person.full_name, line))

            # --- Onomastică (doar tokenuri recunoscute: sfinți ficși + Florii) ---
            if person.mentiuni_tokens:
                matched_saints = []
                for token in person.mentiuni_tokens:
                    norm = _normalize_token(token)
                    if norm == FLORII_TOKEN:
                        occ = next_florii(today)
                    elif norm in FIXED_SAINTS:
                        month, day = FIXED_SAINTS[norm]
                        occ = next_occurrence(month, day, today)
                    else:
                        continue
                    if occ == target_date:
                        matched_saints.append(SAINT_DISPLAY.get(norm, token.strip().title()))
                if matched_saints:
                    saints_txt = ", ".join(matched_saints)
                    line = f"⛪ {person.nume} {person.prenume} — onomastică ({saints_txt})"
                    if person.domiciliu:
                        line += f" | Domiciliu: {person.domiciliu}"
                    nameday_lines.append((person.full_name, line))

        birthday_lines.sort(key=lambda x: x[0])
        nameday_lines.sort(key=lambda x: x[0])
        report.birthdays = [line for _, line in birthday_lines]
        report.namedays = [line for _, line in nameday_lines]

    return report_tomorrow, report_nextweek


# ---------------------------------------------------------------------------
# Formatarea mesajului
# ---------------------------------------------------------------------------


def format_message(report_tomorrow: DayReport, report_nextweek: DayReport) -> str | None:
    if report_tomorrow.is_empty and report_nextweek.is_empty:
        return None

    lines = ["🔔 Zile de naștere & onomastici", ""]

    for report in (report_tomorrow, report_nextweek):
        if report.is_empty:
            continue
        date_str = report.target_date.strftime("%d.%m.%Y")
        lines.append(f"📅 {report.label} ({date_str}):")
        lines.extend(report.birthdays)
        lines.extend(report.namedays)
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


# ---------------------------------------------------------------------------
# Trimitere WhatsApp (Meta Cloud API)
# ---------------------------------------------------------------------------


def send_whatsapp_message(message: str, to_number: str, phone_number_id: str, access_token: str) -> None:
    url = f"https://graph.facebook.com/{WHATSAPP_API_VERSION}/{phone_number_id}/messages"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": to_number,
        "type": "text",
        "text": {"body": message},
    }
    response = requests.post(url, headers=headers, json=payload, timeout=30)
    if response.status_code >= 400:
        print(f"[EROARE] WhatsApp API a răspuns cu {response.status_code}: {response.text}", file=sys.stderr)
        response.raise_for_status()
    print(f"[OK] Mesaj trimis. Răspuns API: {response.status_code} {response.text}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def parse_args():
    parser = argparse.ArgumentParser(
        description="Notificări zile de naștere & onomastici pentru parohie (WhatsApp)."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Doar calculează și printează mesajul, NU trimite nimic pe WhatsApp.",
    )
    parser.add_argument(
        "--date",
        type=str,
        default=None,
        help="Suprascrie data de 'azi' pentru testare (format YYYY-MM-DD).",
    )
    parser.add_argument(
        "--file",
        type=str,
        default=os.environ.get("EXCEL_FILE_PATH", "ZILE.xlsx"),
        help="Calea către fișierul Excel (implicit: ZILE.xlsx din directorul curent).",
    )
    return parser.parse_args()


def main():
    load_dotenv()
    args = parse_args()

    if args.date:
        try:
            today = datetime.datetime.strptime(args.date, "%Y-%m-%d").date()
        except ValueError:
            print(f"Format invalid pentru --date: '{args.date}'. Folosește YYYY-MM-DD.", file=sys.stderr)
            sys.exit(1)
    else:
        today = datetime.date.today()

    try:
        people = read_people(args.file)
    except FileNotFoundError as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)

    print(f"[info] {len(people)} persoane citite cu dată de naștere validă din '{args.file}'.")
    print(f"[info] Data de referință ('azi'): {today.isoformat()}")

    report_tomorrow, report_nextweek = build_reports(people, today)
    message = format_message(report_tomorrow, report_nextweek)

    if message is None:
        print("[info] Nimic de anunțat nici mâine, nici peste o săptămână — nu se trimite niciun mesaj.")
        return

    print("----- MESAJ COMPUS -----")
    print(message)
    print("-------------------------")

    if args.dry_run:
        print("[dry-run] Mesajul NU a fost trimis (mod --dry-run activ).")
        return

    to_number = "40722203265"
    phone_number_id = os.environ.get("META_PHONE_NUMBER_ID")
    access_token = os.environ.get("META_ACCESS_TOKEN")

    if not phone_number_id or not access_token:
        print(
            "[EROARE] Lipsesc META_PHONE_NUMBER_ID și/sau META_ACCESS_TOKEN din mediu "
            "(.env sau GitHub Secrets). Nu pot trimite mesajul.",
            file=sys.stderr,
        )
        sys.exit(1)

    send_whatsapp_message(message, to_number, phone_number_id, access_token)


if __name__ == "__main__":
    main()
